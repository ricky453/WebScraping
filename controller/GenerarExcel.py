import time
from datetime import datetime
from datetime import date
#Para que la bd
import sys
sys.path.insert(1, 'model/operations/')
sys.path.insert(1, 'temp/chromedriver/')
sys.path.insert(1, 'temp/excel/')
from operationCinepolis import OperationCinepolis
#librerias excel
from openpyxl import Workbook
from openpyxl import load_workbook



class GenerarExcel:

    consultaSql = OperationCinepolis()

    def __init__(self):
        sql = "vacio"

    def generarCarteleraExcel(self, fechaBuscar):
        listaCines = self.consultaSql.getCines(fechaBuscar)  #obtiene lista de cines
        
        #creando excel
        wb = Workbook()
        fechaActual = date.today()
        pathSaveExcel = "temp/excel/" + fechaActual.strftime('%d-%m-%Y') + ".xlsx"

        wb.save(pathSaveExcel)  #creando excel
        wb = Workbook()
        wb = load_workbook(pathSaveExcel)  #seleccionando libro
        
        for item in listaCines:  #recorre la lista de cines

            listaPeliculasCine = self.consultaSql.getPeliculasFunciones(item[0], fechaBuscar)  #obtiene peliculas por cine
            #print("mi print cines: " + item[0])
            columnaPelicula = 1  #contadores
            filaPelicula = 1
            filaHora = 1  
            columnaHora = 2

            listaPeliculasCineHoras = self.consultaSql.getPeliculasPorCine(item[0], fechaBuscar) #get horas y peliculas
            ws = wb.create_sheet(item[0],1)  #creando hoja
            wb.save(pathSaveExcel)

            #print(listaPeliculasCineHoras[0][1])

            for elemento in listaPeliculasCine:  #recorre peliculas por cine

                wb = Workbook()
                wb = load_workbook(pathSaveExcel)  #seleccionando libro
                #wsNames.title = "Prueba"  #podemos cambiar nombre de hoja

                ws = wb[item[0]]  #seleccionando hoja (cine)

                ws.cell(row = filaPelicula, column= columnaPelicula).value = elemento[0]  #editando, agregando nombre pelicula
                contadorHora = 0  #contadores

                for itemHoras in listaPeliculasCineHoras:
                    if itemHoras[1] == elemento[0]:
                        ws.cell(row = filaHora, column = columnaHora).value = itemHoras[0]  #editando, agregando las horas
                        columnaHora= columnaHora + 1
                    contadorHora = contadorHora + 1

                filaHora = filaHora + 1
                filaPelicula = filaPelicula + 1
                columnaHora = 2

                wb.save(pathSaveExcel)
                
        print("Excel Generado con exito")







objExcel = GenerarExcel()
objExcel.generarCarteleraExcel(date.today())